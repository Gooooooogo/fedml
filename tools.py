import torch
def average(list):
    transposed = zip(* list)
    averages = [sum(column) / len(column) for column in transposed]
    return averages

def average_for_dict(list_of_dicts):
        sums = {}
        counts = {}

        # 遍历列表中的每个字典
        for d in list_of_dicts:
            for key, value in d.items():
                # 如果键不在 sums 中，将其添加，并将值和计数初始化为0
                sums.setdefault(key, 0)
                counts.setdefault(key, 0)
                # 将值添加到总和
                sums[key] += value
                # 增加计数
                counts[key] += 1
        # 计算平均值
        averages = {key: sums[key] / counts[key] for key in sums}
        return averages


def choose_device():
    if torch.cuda.is_available():
        device = torch.device("cuda")
    else:
        device = torch.device("cpu")
    return device


import openpyxl
def cleanexcel(excel_name, sheet_name):
    workbook = openpyxl.load_workbook(excel_name)

    # 指定要选择的工作表名称
    sheet = workbook[sheet_name]
   # # 删除工作表中的所有旧数据
    sheet.delete_rows(1,sheet.max_row)
    sheet.delete_cols(1,sheet.max_column)
    workbook.save(excel_name)
    workbook.close()

def save2excel(excel_name, sheet_name, data):
    workbook = openpyxl.load_workbook(excel_name)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
    sheet = workbook[sheet_name]

 
    # 获取工作表中的最后一行
    last_row = sheet.max_row + 1
    '''
        data = {
        "current_global_round": current_global_round,
        "current_iteration": current_iteration,
        "average_training_loss": average_training_loss,
        "test_loss": test_loss,
        "accuracy": accuracy
        }
    '''
    column=1
    # 将数据写入选定的工作表，每次一行
    for key, value in data.items():
        sheet.cell(row=last_row, column=column, value=value)  # 写入值
        column+= 1
    # 保存Excel文件
    workbook.save(excel_name)

    # 关闭工作簿
    workbook.close()

def save2excel_batch(excel_name, sheet_name, data):
    workbook = openpyxl.load_workbook(excel_name)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
    sheet = workbook[sheet_name]

 
    # 获取工作表中的最后一行
    last_row = sheet.max_row + 1
    '''
        data = {
        "current_global_round": current_global_round,
        "current_iteration": current_iteration,
        "average_training_loss": average_training_loss,
        "test_loss": test_loss,
        "accuracy": accuracy
        }
    '''
    
    # 将数据写入选定的工作表，每次一行
    for dt in data:  
        column=1
        for key, value in dt.items():
            sheet.cell(row=last_row, column=column, value=value)  # 写入值
            column+= 1
        last_row +=1 
    # 保存Excel文件
    workbook.save(excel_name)

    # 关闭工作簿
    workbook.close()

def set_client(client_num):
    client_list=[]
    for i in range(client_num):
        client_list.append('client'+str(i+1))
    return client_list